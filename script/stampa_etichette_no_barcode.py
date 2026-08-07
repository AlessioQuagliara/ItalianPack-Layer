#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
from reportlab.lib.pagesizes import mm
from reportlab.pdfgen import canvas
from openpyxl import load_workbook

def ask_file():
    return input("Inserisci il percorso del file Excel (es. etichette.xlsx): ").strip()

def read_excel_data(filepath):
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active
    
    # Cerchiamo la colonna "Codice", altrimenti prendiamo di default la prima colonna (indice 0)
    headers = [cell.value for cell in ws[1]]
    codice_col = 0
    for idx, h in enumerate(headers):
        if h and isinstance(h, str):
            if h.strip().lower() == "codice":
                codice_col = idx
                break

    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[codice_col] is not None:
            data.append(str(row[codice_col]))
    return data

def generate_pdf(data, output_filename="etichette.pdf", page_width_mm=70, page_height_mm=27):
    # --- CONFIGURAZIONE ---
    page_width_pt = page_width_mm * mm
    page_height_pt = page_height_mm * mm
    
    c = canvas.Canvas(output_filename, pagesize=(page_width_pt, page_height_pt))
    
    # Font configurabile per dimensione
    font_name = "Helvetica"
    font_size = 14  # Puoi aumentare o diminuire questo valore per la grandezza del testo
    
    for idx, testo in enumerate(data):
        if idx > 0: 
            c.showPage()
        
        c.setFont(font_name, font_size)
        
        # Calcolo della larghezza del testo per centratura orizzontale
        text_width = c.stringWidth(testo, font_name, font_size)
        text_x = (page_width_pt - text_width) / 2
        
        # Calcolo dell'altezza per centratura verticale 
        # (font_size * 0.35 approssima la distanza della baseline dal centro geometrico del testo)
        text_y = (page_height_pt / 2) - (font_size * 0.35)
        
        # Stampa del testo centrale
        c.drawString(text_x, text_y, testo)

    c.save()
    print(f"PDF generato: {output_filename}")

if __name__ == "__main__":
    filepath = ask_file()
    if os.path.isfile(filepath):
        generate_pdf(read_excel_data(filepath))
    else:
        print("File non trovato.")