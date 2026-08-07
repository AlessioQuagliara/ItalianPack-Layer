#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
from reportlab.lib.pagesizes import mm
from reportlab.pdfgen import canvas
from reportlab.graphics.barcode import code39
from openpyxl import load_workbook
from textwrap import wrap
from decimal import Decimal

def ask_file():
    return input("Inserisci il percorso del file Excel (es. etichette.xlsx): ").strip()

def read_excel_data(filepath):
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active
    
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_val = row[0]
        
        # 1. GESTIONE BARCODE (COLONNA 1)
        if raw_val is None:
            barcode_val = ""
        elif isinstance(raw_val, (float, int)):
            # Usiamo Decimal per preservare l'esattezza dei numeri lunghi di Excel
            barcode_val = "{:f}".format(Decimal(repr(raw_val))).split('.')[0]
        else:
            barcode_val = str(raw_val).strip()

        # --- CORREZIONE PDA ---
        # Se è puramente numerico e lungo 19, tagliamo gli ultimi 2 zeri.
        # Se è alfanumerico (es. con la 'H'), lo lasciamo intatto.
        if barcode_val.isdigit() and len(barcode_val) == 19:
            barcode_val = barcode_val[:-2]

        # 2. ALTRE COLONNE
        descrizione = str(row[1]) if len(row) > 1 and row[1] is not None else ""
        codice_display = str(row[2]) if len(row) > 2 and row[2] is not None else ""
        
        if barcode_val or descrizione or codice_display:
            data.append((barcode_val, descrizione, codice_display))
    return data

def sanitize_code39(codice):
    allowed = set("ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789-. $/+%")
    sanitized = ''.join(c if c in allowed else ' ' for c in codice.upper())
    return sanitized.strip() or "0"

def generate_pdf(data, output_filename="etichette_speciali.pdf", page_width_mm=70, page_height_mm=27):
    # --- CONFIGURAZIONE ---
    margin_side_pt = 2 * mm
    page_width_pt = page_width_mm * mm
    page_height_pt = page_height_mm * mm
    usable_width_pt = page_width_pt - (2 * margin_side_pt)
    
    c = canvas.Canvas(output_filename, pagesize=(page_width_pt, page_height_pt))
    
    for idx, (b_val, desc, c_disp) in enumerate(data):
        if idx > 0: c.showPage()
        
        # 1. BARCODE (Dalla prima colonna, eventualmente tagliato a 17)
        sanitized_b = sanitize_code39(b_val)
        bw = 0.5 * mm 
        barcode_h = 8 * mm
        barcode = code39.Standard39(
            sanitized_b, 
            barWidth=bw, 
            barHeight=barcode_h, 
            humanReadable=False, 
            checksum=False
        )
        
        if barcode.width > usable_width_pt:
            barcode.barWidth = bw * (usable_width_pt / barcode.width)
        
        bx = margin_side_pt + (usable_width_pt - barcode.width) / 2
        by = page_height_pt - barcode.barHeight 
        barcode.drawOn(c, bx, by)
        
        # 2. CODICE DISPLAY (Dalla terza colonna - 20pt Bold)
        # Abbassato leggermente di più per non toccare le barre
        current_y = by - 8 * mm 
        font_size_code = 20
        c.setFont("Helvetica-Bold", font_size_code)
        
        c_disp_width = c.stringWidth(c_disp, "Helvetica-Bold", font_size_code)
        # Se il testo gigante esce dai bordi, lo ridimensioniamo al volo
        if c_disp_width > usable_width_pt:
            font_size_code = font_size_code * (usable_width_pt / c_disp_width)
            c.setFont("Helvetica-Bold", font_size_code)
            c_disp_width = c.stringWidth(c_disp, "Helvetica-Bold", font_size_code)
            
        cx = margin_side_pt + (usable_width_pt - c_disp_width) / 2
        c.drawString(cx, current_y, c_disp)
        
        # 3. DESCRIZIONE (Dalla seconda colonna)
        # Calcoliamo lo spazio basandoci sull'altezza del font grande (circa 7mm)
        current_y -= 6 * mm 
        font_size_desc = 8
        c.setFont("Helvetica", font_size_desc)
        
        lines = wrap(desc, width=45)
        line_h = font_size_desc + 1.5
        
        for line in lines:
            if current_y > 1.5 * mm: # Margine di sicurezza basso
                tw = c.stringWidth(line, "Helvetica", font_size_desc)
                tx = margin_side_pt + (usable_width_pt - tw) / 2
                c.drawString(tx, current_y, line)
                current_y -= line_h

    c.save()
    print(f"PDF Generato: {os.path.abspath(output_filename)}")

if __name__ == "__main__":
    path = ask_file()
    if os.path.isfile(path):
        generate_pdf(read_excel_data(path))
    else:
        print("File non trovato.")