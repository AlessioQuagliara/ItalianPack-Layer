#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
from reportlab.lib.pagesizes import mm
from reportlab.pdfgen import canvas
from reportlab.graphics.barcode import code39
from openpyxl import load_workbook
from textwrap import wrap

def ask_file():
    return input("Inserisci il percorso del file Excel (es. etichette.xlsx): ").strip()

def read_excel_data(filepath):
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    
    codice_col, descrizione_col = 0, 1
    for idx, h in enumerate(headers):
        if h and isinstance(h, str):
            h_low = h.strip().lower()
            if h_low == "codice": codice_col = idx
            elif h_low == "descrizione": descrizione_col = idx

    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[codice_col] is None and row[descrizione_col] is None: continue
        codice = str(row[codice_col]) if row[codice_col] is not None else ""
        descrizione = str(row[descrizione_col]) if row[descrizione_col] is not None else ""
        data.append((codice, descrizione))
    return data

def sanitize_code39(codice):
    allowed = set("ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789-. $/+%")
    sanitized = ''.join(c if c in allowed else ' ' for c in codice.upper())
    return sanitized.strip() or "CODICE"

def generate_pdf(data, output_filename="etichette.pdf", page_width_mm=70, page_height_mm=27):
    # --- CONFIGURAZIONE ---
    margin_side_pt = 2 * mm  # Margine laterale mantenuto
    page_width_pt = page_width_mm * mm
    page_height_pt = page_height_mm * mm
    usable_width_pt = page_width_pt - (2 * margin_side_pt)
    
    c = canvas.Canvas(output_filename, pagesize=(page_width_pt, page_height_pt))
    
    for idx, (codice, descrizione) in enumerate(data):
        if idx > 0: c.showPage()
        
        sanitized = sanitize_code39(codice)
        
        # 1. DISEGNO BARCODE (Attaccato al bordo superiore)
        bw = 0.4 * mm 
        barcode_height = 8 * mm
        barcode = code39.Standard39(
            sanitized, 
            barWidth=bw, 
            barHeight=barcode_height, 
            humanReadable=False,
            checksum=False # Forza la disattivazione del carattere di controllo
        )
        
        # Adattamento larghezza se necessario
        if barcode.width > usable_width_pt:
            barcode.barWidth = bw * (usable_width_pt / barcode.width)
        
        barcode_x = margin_side_pt + (usable_width_pt - barcode.width) / 2
        barcode_y = page_height_pt - barcode.barHeight 
        
        barcode.drawOn(c, barcode_x, barcode_y)
        
        # 2. DISEGNO TESTO CODICE (subito sotto il barcode)
        # Modificato: rimosso Bold e aumentata dimensione a 10
        current_y = barcode_y - 4 * mm  
        font_name_code = "Helvetica"
        font_size_code = 10
        
        c.setFont(font_name_code, font_size_code)
        code_text_width = c.stringWidth(sanitized, font_name_code, font_size_code)
        c.drawString(margin_side_pt + (usable_width_pt - code_text_width) / 2, current_y, sanitized)
        
        # 3. DISEGNO DESCRIZIONE
        current_y -= 5 * mm 
        
        font_size_desc = 8
        c.setFont("Helvetica", font_size_desc)
        
        # Wrap del testo
        wrapper_width = 42 
        lines = wrap(descrizione, width=wrapper_width)
        
        line_height = font_size_desc + 1.5
        
        for line in lines:
            if current_y > 2 * mm: 
                text_width = c.stringWidth(line, "Helvetica", font_size_desc)
                text_x = margin_side_pt + (usable_width_pt - text_width) / 2
                c.drawString(text_x, current_y, line)
                current_y -= line_height

    c.save()
    print(f"PDF generato: {output_filename}")

if __name__ == "__main__":
    filepath = ask_file()
    if os.path.isfile(filepath):
        generate_pdf(read_excel_data(filepath))
    else:
        print("File non trovato.")