from openpyxl import load_workbook

COLUMN_ALIASES = {
    "codice": "codice",
    "descrizione": "descrizione",
    "cassetto": "cassetto",
    "ubicazione": "ubicazione",
    "quantita": "quantita",
    "quantità": "quantita",
    "mancanti": "mancanti",
    "rda": "rda",
    "avanzi": "avanzi",
    "gruppo": "gruppo",
    "descrizione gruppo": "descrizione_gruppo",
}

NUMERIC_FIELDS = {"quantita", "mancanti", "avanzi"}


def _normalize_header(value):
    if value is None:
        return None
    return " ".join(str(value).strip().lower().split())


def _to_float(value):
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def read_distinta(filepath):
    """Legge la distinta Excel e restituisce una lista di dict pronti per RigaCommessa."""
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    headers = [_normalize_header(cell.value) for cell in ws[1]]
    field_by_col = {}
    for idx, header in enumerate(headers):
        field = COLUMN_ALIASES.get(header)
        if field:
            field_by_col[idx] = field

    if "codice" not in field_by_col.values():
        raise ValueError("Il file Excel non contiene una colonna 'Codice'.")

    righe = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = {}
        for idx, field in field_by_col.items():
            value = row[idx] if idx < len(row) else None
            if field in NUMERIC_FIELDS:
                value = _to_float(value)
            elif value is not None:
                value = str(value).strip()
            record[field] = value

        if not record.get("codice"):
            continue

        righe.append(record)

    if not righe:
        raise ValueError("Nessuna riga valida trovata nel file Excel (manca il Codice).")

    return righe
